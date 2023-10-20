import pyodbc 
import pandas as pd
import numpy as np
from sklearn import datasets
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
import time
import sys
from openpyxl import Workbook
import matplotlib.patches as mpatches
from scipy import stats
import openpyxl
import streamlit as st

from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV, StratifiedKFold
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.metrics import accuracy_score, confusion_matrix, ConfusionMatrixDisplay, classification_report
from sklearn.metrics import roc_curve, roc_auc_score, precision_score, recall_score, f1_score, RocCurveDisplay
import requests
from io import StringIO
import gdown

# orig_url='https://drive.google.com/file/d/1i3Oe4429eUDRJiBH_OP9Qb7gSsW1p5MH/view?usp=sharing'
# down_url='https://drive.google.com/uc?id=' + orig_url.split('/')[-2]
# #df = pd.read_csv(down_url)
# df = pd.read_csv(down_url, encoding='utf8')

df = pd.read_csv('data/data_process.csv')
samples, features = df.shape
print('Number Of Samples: ', samples) # how many rows/columns
print('Number Of Features: ', features) # how many rows/columns
print(df.head())
print(df.info())
print(df.describe().T)
print(df.tail())
d = []
u = []
t = []
for col in df:
    d.append(col)
    u.append(df[col].nunique())
    t.append(df[col].dtype)
print(pd.DataFrame({'column':d,'type': t ,'unique value' : u}))
print("_________________________________________________________")

#Preprocessing and Cleaning Data
print(df.isnull().sum())
print(df.duplicated().sum())
print(df.info())
print(df.shape)


labels = ['ChuaXemSMS', 'DaXemSMS'] # define gender
values = df["TrangThaiXem"].value_counts().values # count TrangThaiXem
df["TrangThaiXem"].replace({True: 'DaXemSMS', False: 'ChuaXemSMS'}, inplace=True)
plt.style.use('fivethirtyeight')
plt.figure(figsize=(12, 8))
plt.subplot(1, 2, 1)
plt.title('Total Status Open SMS/ Not Open SMS')
sns.countplot(x=df["TrangThaiXem"], data=df)
plt.subplot(1, 2, 2)
plt.pie(values, labels=labels, autopct='%1.1f%%')
plt.savefig('viewsms')
plt.show()

top10 = df.groupby("VanPhong").agg(total_Revenue=("TongTienBL", "sum")).nlargest(10, "total_Revenue").reset_index()
print(top10)
fix, ax = plt.subplots(figsize=(18,8))
sns.barplot(x="total_Revenue", y = "VanPhong", dodge=False, orient="h", data = top10)
max_value = top10['total_Revenue'].max()
ax.set_xlim([0,max_value])
ax.set_xticklabels(['{:,.0f}'.format(x) for x in ax.get_xticks()])
ax.set_ylabel("VanPhong")
ax.set_title('Top 10 Revenue Employee')
plt.savefig('top10_revenue')
plt.show()

top10.to_excel("top10.xlsx", index= False)
wb = openpyxl.load_workbook("top10.xlsx")
wb.worksheets[0].title = "data"
chart = wb[wb.sheetnames[0]]
img = openpyxl.drawing.image.Image("top10_revenue.png")
img.anchor = "C2"
chart.add_image(img)
chart.cell(row = 1, column = 3).value = "Top 10 Revenue Employee"
wb.save("top10.xlsx")

top10_cb = df.groupby(['VanPhong', 'UserLayMau']).agg(total_Revenue=("TongTienBL", "sum")).nlargest(10, "total_Revenue").reset_index()
# data={
#         'Region':['Central', 'East', 'South', 'West'],
#         'States':['Illinois', 'New York', 'Florida', 'California'],
#         'Sales':[98971.25, 223930.48, 87651.11, 288310.61],
#     }
# df = pd.DataFrame(data)

top10_cb = top10_cb.sort_values('total_Revenue', ascending=False)
fix, ax = plt.subplots(figsize=(24,16))
ax.bar(top10_cb['UserLayMau'], top10_cb['total_Revenue'], label=top10_cb['VanPhong'], color=['orange', 'green', 'pink', 'grey'])
ax.set_xticklabels(top10_cb['UserLayMau'], fontsize=12)
max_value = top10_cb['total_Revenue'].max()
ax.set_ylim([0,max_value])
ax.set_yticklabels(['{:,.0f}'.format(x) for x in ax.get_yticks()])
ax.set_title('Top 10 Revenue Employee by Units')
for bar, state in zip(ax.patches, top10_cb['VanPhong']):
    ax.text(bar.get_x()+bar.get_width()/2, 10000, state, rotation=-90, fontsize = 12, color = 'black', ha = 'center', va = 'bottom')
plt.savefig('top10_revenue_byunits')
plt.show()

top10_cb.to_excel("top10_cb.xlsx", index= False)
wb = openpyxl.load_workbook("top10_cb.xlsx")
wb.worksheets[0].title = "data"
chart = wb[wb.sheetnames[0]]
img = openpyxl.drawing.image.Image("top10_revenue_byunits.png")
img.anchor = "D2"
chart.add_image(img)
chart.cell(row = 1, column = 4).value = "Top 10 Revenue Employee by Units" #This will change the cell(2,4) to 4
wb.save("top10_cb.xlsx")

# write mutil table on a sheet 
with pd.ExcelWriter('top10_sumary.xlsx', engine='xlsxwriter') as writer:
    top10_cb.to_excel(writer, sheet_name='data', startrow=1, startcol=0)
    top10.to_excel(writer, sheet_name='data', startrow=1+len(top10_cb)+3, startcol=0) 


daxem = df.groupby('VanPhong')['TrangThaiXem'].sum().reset_index()
print(daxem)
chuaxem = df[df["TrangThaiXem"]==0].groupby('VanPhong')['TrangThaiXem'].count().reset_index()
print(chuaxem)
fix, ax = plt.subplots(figsize=(16,10))
ax.bar(daxem['VanPhong'], daxem["TrangThaiXem"], label='DaxemSMS')
ax.bar(chuaxem['VanPhong'], chuaxem["TrangThaiXem"], bottom=daxem['TrangThaiXem'],
       label='ChuaXemSMS')
ax.set_title('Total Open SMS by Units')
xticks = sorted(daxem['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90, fontsize = 10)
ax.legend()
plt.savefig('totalstatussmsbyunits')
plt.show()

data={
       'VanPhong':daxem['VanPhong'],
       'DaxemSMS': daxem["TrangThaiXem"],
       'ChuaXemSMS':chuaxem["TrangThaiXem"]
    }
df_frame = pd.DataFrame(data)
df_frame.to_excel("viewsms.xlsx", index= False)
wb = openpyxl.load_workbook("viewsms.xlsx")
wb.worksheets[0].title = "data"
chart = wb[wb.sheetnames[0]]
img = openpyxl.drawing.image.Image("viewsms.png")
img.anchor = "D2"
chart.add_image(img)
chart.cell(row = 1, column = 4).value = "Total Status Open SMS/ Not Open SMS"
wb.save("viewsms.xlsx")

DaTuVan = df.groupby('VanPhong')['TuVan'].sum().reset_index()
ChuaTuVan = df[df["TuVan"]==0].groupby('VanPhong')['TuVan'].count().reset_index()
fix, ax = plt.subplots(figsize=(16,10))
ax.bar(DaTuVan['VanPhong'], DaTuVan["TuVan"], label='DaTuVan')
ax.bar(ChuaTuVan['VanPhong'], ChuaTuVan["TuVan"], bottom=DaTuVan['TuVan'],
       label='ChuaTuVan')
ax.set_title('Total Consultant by Units')
xticks = sorted(DaTuVan['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90, fontsize = 10)
ax.legend()
plt.savefig('totalconsultantbyunits')
plt.show()

data={
       'VanPhong':DaTuVan["VanPhong"],
       'DaTuVan': DaTuVan["TuVan"],
       'ChuaTuVan':ChuaTuVan["TuVan"]
    }
df_frame2 = pd.DataFrame(data)
df_frame2.to_excel("totalconsultantbyunits.xlsx", index= False)
wb = openpyxl.load_workbook("totalconsultantbyunits.xlsx")
wb.worksheets[0].title = "data"
chart = wb[wb.sheetnames[0]]
img = openpyxl.drawing.image.Image("totalconsultantbyunits.png")
img.anchor = "D2"
chart.add_image(img)
chart.cell(row = 1, column = 4).value = "Total Consultant by Units"
wb.save("totalconsultantbyunits.xlsx")

#Merge, join, concatenate dataframe
frames = [df_frame, df_frame2]
result = pd.concat(frames)
print(result)

# write mutil table on a sheet 
with pd.ExcelWriter('consultant_viewsms.xlsx', engine='xlsxwriter') as writer:
    df_frame2.to_excel(writer, sheet_name='data', startrow=1, startcol=0)
    df_frame.to_excel(writer, sheet_name='data', startrow=1+len(df_frame2)+3, startcol=0) 
    df_frame2.to_excel(writer, sheet_name='data2', startrow=1, startcol=0)

# read data from mutil sheet 
sheetname_list = ['data', 'data2']
xl = pd.read_excel('consultant_viewsms.xlsx', sheet_name=sheetname_list)
data_1 = xl['data'].values
print(data_1)
data_2 = xl['data2'].values
print(data_2)


#show list columns
print(df.columns.to_list())
#check null
print(df.isnull().sum())
#check type
print(df.dtypes)

#get min/max TongTienBL
print(df["TongTienBL"].agg(["min", "max"]))
print(df.groupby("VanPhong").agg(n_orders=("TongTienBL", "nunique")).sort_index)
df2 = df.groupby(['VanPhong'])['TongTienBL'].sum().reset_index().sort_index()
print(df2)
df1 = df.groupby(['VanPhong'])['SID'].count().reset_index().sort_index()
print(df1)
df3= df.groupby(['VanPhong'])['SLMau'].count().reset_index().sort_index()
print(df3)
df4 = df.groupby('VanPhong')['TrangThaiXem'].sum().reset_index().sort_index()
print(df4)

print(df["SoSeri"].str.split("-").str[1])

fix, ax = plt.subplots(figsize=(16,10))
sns.barplot(x="VanPhong", y = "TongTienBL", data = df2, color = "g")
ax.set_xlabel("VanPhong")
max_value = df2['TongTienBL'].max()
ax.set_ylim([0,max_value])
ax.set_yticklabels(['{:,.0f}'.format(x) for x in ax.get_yticks()])
ax.set_title("Total Revenue by Units")
xticks = sorted(df2['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90)
plt.savefig('totalrevenuebyunits')
plt.show()

df2.to_excel("totalrevenuebyunits.xlsx", index= False)
wb = openpyxl.load_workbook("totalrevenuebyunits.xlsx")
wb.worksheets[0].title = "data"
chart = wb[wb.sheetnames[0]]
img = openpyxl.drawing.image.Image("totalrevenuebyunits.png")
img.anchor = "C2"
chart.add_image(img)
chart.cell(row = 1, column = 3).value = "Total Revenue by Units"
wb.save("totalrevenuebyunits.xlsx")

fix, ax = plt.subplots(figsize=(16,10))
sns.barplot(x="VanPhong", y = "SID", data = df1, color = "r")
ax.set_xlabel("VanPhong")
ax.set_ylabel("Schedule",)
ax.set_title("Total Schedule by Units")
xticks = sorted(df1['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90)
plt.savefig('totalschedulebyunits')
plt.show()

fix, ax = plt.subplots(figsize=(16,10))
sns.barplot(x="VanPhong", y = "SLMau", data = df3, color = "b")
ax.set_xlabel("VanPhong")
ax.set_ylabel("SL Mau",)
ax.set_title("Total Mau by Units")
xticks = sorted(df3['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90)
plt.savefig('totalmaubyunits')
plt.show()

fix, ax = plt.subplots(figsize=(16,10))
sns.barplot(x="VanPhong", y = "TrangThaiXem", data = df4, color='steelblue')
ax.set_xlabel("VanPhong")
ax.set_ylabel("SL Xem",)
ax.set_title("Total Open SMS by Units")
xticks = sorted(df4['VanPhong'].unique())
ax.set_xticklabels(xticks,rotation = -90)
plt.savefig('totalopensmsbyunits')
plt.show()

#-----------------------------------------------------------------------------------
#Exploring Categorical Features
categorical_columns = ['TrangThaiXem', 'TuVan', 'TrangThaiGui', 'TrangThaiSMSKQ']
fig, axes = plt.subplots(4,2, figsize=(20,30))

sns.set_style('darkgrid')
idx = 0
for col in categorical_columns:
    sns.countplot(data=df, y=col, palette='magma', orient='h',
                  ax=axes[idx][0]).set_title(f'Count of {col}', fontsize='10')
    for container in axes[idx][0].containers:
        axes[idx][0].bar_label(container)

    sns.countplot(data=df, y=col, palette='mako', orient='h',
                  ax=axes[idx][1]).set_title(f'Count of {col} per TongTienBL', fontsize='10')
    for container in axes[idx][1].containers:
        axes[idx][1].bar_label(container)
        
    idx +=1
plt.savefig('a2')
plt.show()